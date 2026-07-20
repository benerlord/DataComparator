import json
from pathlib import Path
import pytest
from openpyxl import Workbook
from datacompare.config.loader import load_task_or_batch
from datacompare.config.models import BatchConfig
from datacompare.runner import execute_batch
from datacompare.engine.result import BatchResult


def _make_xlsx(path: Path, rows):
    wb = Workbook(); ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(path)


def _write(path: Path, text: str):
    path.write_text(text, encoding="utf-8")


def _batch_two_success(tmp_path: Path) -> Path:
    _make_xlsx(tmp_path / "left.xlsx", [
        ["order_id", "amount"], ["A1", "1.00"], ["A2", "2.00"],
    ])
    _make_xlsx(tmp_path / "right.xlsx", [
        ["order_id", "amount"], ["A1", "1.00"], ["A2", "2.00"],
    ])
    task = tmp_path / "batch.yaml"
    _write(task, f"""
name: b
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/right.xlsx}}
match:
  keys: [{{left: order_id, right: order_id}}]
compare:
  fields: [{{left: amount, right: amount, mode: numeric, decimal_places: 2}}]
output:
  dir: {tmp_path}/out
  formats: [json]
tasks:
  - name: sub1
  - name: sub2
""")
    return task


def test_execute_batch_all_success(tmp_path):
    task_path = _batch_two_success(tmp_path)
    batch = load_task_or_batch(task_path, {})
    result = execute_batch(batch, connections={})
    assert isinstance(result, BatchResult)
    assert result.success_count == 2
    assert result.failed_count == 0
    assert (tmp_path / "out" / "sub1" / "report.json").exists()
    assert (tmp_path / "out" / "sub2" / "report.json").exists()


def test_execute_batch_continues_on_sub_task_failure(tmp_path):
    _make_xlsx(tmp_path / "left.xlsx", [
        ["order_id", "amount"], ["A1", "1.00"],
    ])
    _make_xlsx(tmp_path / "right.xlsx", [
        ["order_id", "amount"], ["A1", "1.00"],
    ])
    task = tmp_path / "batch.yaml"
    _write(task, f"""
name: b
on_error: continue
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/right.xlsx}}
match:
  keys: [{{left: order_id, right: order_id}}]
compare:
  fields: [{{left: amount, right: amount, mode: numeric, decimal_places: 2}}]
output:
  dir: {tmp_path}/out
  formats: [json]
tasks:
  - name: sub1
  - name: sub2
    sources:
      left: {{path: {tmp_path}/missing.xlsx}}
""")
    batch = load_task_or_batch(task, {})
    result = execute_batch(batch, connections={})
    assert result.success_count == 1
    assert result.failed_count == 1
    assert result.task_results[0].status == "success"
    assert result.task_results[1].status == "failed"
    assert result.task_results[1].error is not None


def test_execute_batch_sub_task_explicit_output_dir_overrides_autopath(tmp_path):
    _make_xlsx(tmp_path / "left.xlsx", [["order_id"], ["A1"]])
    _make_xlsx(tmp_path / "right.xlsx", [["order_id"], ["A1"]])
    custom_dir = tmp_path / "custom_place"
    task = tmp_path / "batch.yaml"
    _write(task, f"""
name: b
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/right.xlsx}}
match:
  keys: [{{left: order_id, right: order_id}}]
compare:
  fields: []
output:
  dir: {tmp_path}/out
  formats: [json]
tasks:
  - name: sub_custom
    output:
      dir: {custom_dir}
""")
    batch = load_task_or_batch(task, {})
    execute_batch(batch, connections={})
    assert (custom_dir / "report.json").exists()
    # auto-path NOT used
    assert not (tmp_path / "out" / "sub_custom").exists() or \
        not (tmp_path / "out" / "sub_custom" / "report.json").exists()


def test_execute_batch_fail_fast_skips_remaining(tmp_path):
    _make_xlsx(tmp_path / "left.xlsx", [["order_id"], ["A1"]])
    _make_xlsx(tmp_path / "right.xlsx", [["order_id"], ["A1"]])
    task = tmp_path / "batch.yaml"
    _write(task, f"""
name: b
on_error: fail_fast
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/right.xlsx}}
match:
  keys: [{{left: order_id, right: order_id}}]
compare:
  fields: []
output:
  dir: {tmp_path}/out
  formats: [json]
tasks:
  - name: ok1
  - name: broken
    sources: {{left: {{path: {tmp_path}/missing.xlsx}}}}
  - name: never_runs
  - name: also_skipped
""")
    batch = load_task_or_batch(task, {})
    result = execute_batch(batch, connections={})
    assert result.success_count == 1
    assert result.failed_count == 1
    assert result.skipped_count == 2
    assert result.task_results[0].status == "success"
    assert result.task_results[1].status == "failed"
    assert result.task_results[2].status == "skipped"
    assert result.task_results[3].status == "skipped"
    # skipped sub-tasks should NOT have created output dirs
    assert not (tmp_path / "out" / "never_runs").exists()
    assert not (tmp_path / "out" / "also_skipped").exists()


def test_execute_batch_fail_fast_reports_zero_duration_for_skipped(tmp_path):
    _make_xlsx(tmp_path / "left.xlsx", [["order_id"], ["A1"]])
    task = tmp_path / "batch.yaml"
    _write(task, f"""
name: b
on_error: fail_fast
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/left.xlsx}}
match:
  keys: [{{left: order_id, right: order_id}}]
compare:
  fields: []
output:
  dir: {tmp_path}/out
  formats: [json]
tasks:
  - name: fail_me
    sources: {{left: {{path: {tmp_path}/nope.xlsx}}}}
  - name: skipped_task
""")
    batch = load_task_or_batch(task, {})
    result = execute_batch(batch, connections={})
    assert result.task_results[1].status == "skipped"
    assert result.task_results[1].duration_ms == 0


def _read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def test_execute_batch_writes_batch_log_with_start_and_end(tmp_path):
    task = _batch_two_success(tmp_path)
    batch = load_task_or_batch(task, {})
    execute_batch(batch, connections={})
    batch_log = tmp_path / "out" / "batch.log"
    assert batch_log.exists()
    entries = _read_jsonl(batch_log)
    events = [e["event"] for e in entries]
    assert events[0] == "batch_start"
    assert events[-1] == "batch_end"
    assert events.count("task_start") == 2
    assert events.count("task_end") == 2


def test_batch_log_task_end_carries_status_and_counts(tmp_path):
    task = _batch_two_success(tmp_path)
    batch = load_task_or_batch(task, {})
    execute_batch(batch, connections={})
    entries = _read_jsonl(tmp_path / "out" / "batch.log")
    task_ends = [e for e in entries if e["event"] == "task_end"]
    assert all(e["status"] == "success" for e in task_ends)
    assert all("matched" in e and "diff" in e for e in task_ends)
    assert all(isinstance(e["duration_ms"], int) for e in task_ends)


def test_batch_log_records_failure_with_error_type_and_message(tmp_path):
    _make_xlsx(tmp_path / "left.xlsx", [["order_id"], ["A1"]])
    _make_xlsx(tmp_path / "right.xlsx", [["order_id"], ["A1"]])
    task = tmp_path / "batch.yaml"
    _write(task, f"""
name: b
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/right.xlsx}}
match:
  keys: [{{left: order_id, right: order_id}}]
compare:
  fields: []
output:
  dir: {tmp_path}/out
  formats: [json]
tasks:
  - name: broken
    sources: {{left: {{path: {tmp_path}/missing.xlsx}}}}
""")
    batch = load_task_or_batch(task, {})
    execute_batch(batch, connections={})
    entries = _read_jsonl(tmp_path / "out" / "batch.log")
    task_end = next(e for e in entries if e["event"] == "task_end")
    assert task_end["status"] == "failed"
    assert "error_type" in task_end
    assert "error_message" in task_end


def test_batch_log_end_event_has_final_counts(tmp_path):
    task = _batch_two_success(tmp_path)
    batch = load_task_or_batch(task, {})
    execute_batch(batch, connections={})
    entries = _read_jsonl(tmp_path / "out" / "batch.log")
    end = next(e for e in entries if e["event"] == "batch_end")
    assert end["success"] == 2
    assert end["failed"] == 0
    assert end["skipped"] == 0
    assert isinstance(end["total_duration_ms"], int)


# ---------------------------------------------------------------------------
# Task 3: batch_summary.{json,html} + fail_on_diff-aware exit_code
# ---------------------------------------------------------------------------

def _make_xlsx_multi(path: Path, sheets: dict):
    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "_placeholder"
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    if "_placeholder" in wb.sheetnames:
        del wb["_placeholder"]
    wb.save(path)


def test_execute_batch_writes_batch_summary_json_and_html(tmp_path):
    """execute_batch produces batch_summary.{json,html} in the aggregate out_dir."""
    _make_xlsx_multi(tmp_path / "l.xlsx", {"S": [["id"], ["x"]]})
    _make_xlsx_multi(tmp_path / "r.xlsx", {"S": [["id"], ["x"]]})
    yaml_path = tmp_path / "batch.yaml"
    yaml_path.write_text(f"""
name: agg_test
sources:
  left: {{type: excel, path: {tmp_path}/l.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: t1
    sources:
      left: {{sheets: [{{name: S}}]}}
      right: {{type: excel, path: {tmp_path}/r.xlsx, sheets: [{{name: S}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
""", encoding="utf-8")
    cfg = load_task_or_batch(yaml_path)
    assert isinstance(cfg, BatchConfig)
    execute_batch(cfg, {}, fail_on_diff=False)
    summary_json = tmp_path / "reports" / "batch_summary.json"
    summary_html = tmp_path / "reports" / "batch_summary.html"
    assert summary_json.exists()
    assert summary_html.exists()
    data = json.loads(summary_json.read_text(encoding="utf-8"))
    assert data["batch_name"] == "agg_test"
    assert data["task_count"] == 1
    assert data["success_count"] == 1
    assert data["exit_code"] == 0
    assert data["tasks"][0]["name"] == "t1"
    assert data["tasks"][0]["report_dir"] == "t1"


def test_execute_batch_summary_exit_code_reflects_fail_on_diff(tmp_path):
    """With diffs and fail_on_diff=True, summary exit_code should be 10."""
    _make_xlsx_multi(tmp_path / "l.xlsx", {"S": [["id", "v"], ["x", "1"]]})
    _make_xlsx_multi(tmp_path / "r.xlsx", {"S": [["id", "v"], ["x", "2"]]})
    yaml_path = tmp_path / "batch.yaml"
    yaml_path.write_text(f"""
name: diff_test
sources:
  left: {{type: excel, path: {tmp_path}/l.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: t1
    sources:
      left: {{sheets: [{{name: S}}]}}
      right: {{type: excel, path: {tmp_path}/r.xlsx, sheets: [{{name: S}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: v, right: v}}]}}
""", encoding="utf-8")
    cfg = load_task_or_batch(yaml_path)
    execute_batch(cfg, {}, fail_on_diff=True)
    data = json.loads(
        (tmp_path / "reports" / "batch_summary.json").read_text(encoding="utf-8")
    )
    assert data["exit_code"] == 10  # diff + fail_on_diff


def test_execute_batch_summary_records_failed_task(tmp_path):
    """A sub-task that raises a non-ConfigError should appear with status=failed,
    error info, and cause the summary exit_code to be 2 (runtime error)."""
    _make_xlsx_multi(tmp_path / "l.xlsx", {"S": [["id"], ["x"]]})
    _make_xlsx_multi(tmp_path / "r.xlsx", {"S": [["id"], ["x"]]})
    yaml_path = tmp_path / "batch.yaml"
    yaml_path.write_text(f"""
name: fail_test
on_error: continue
sources:
  left: {{type: excel, path: {tmp_path}/l.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: t_ok
    sources:
      left: {{sheets: [{{name: S}}]}}
      right: {{type: excel, path: {tmp_path}/r.xlsx, sheets: [{{name: S}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
  - name: t_bad_file
    sources:
      left: {{sheets: [{{name: S}}]}}
      right: {{type: excel, path: {tmp_path}/does_not_exist.xlsx, sheets: [{{name: S}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
""", encoding="utf-8")
    cfg = load_task_or_batch(yaml_path)
    execute_batch(cfg, {}, fail_on_diff=False)
    data = json.loads(
        (tmp_path / "reports" / "batch_summary.json").read_text(encoding="utf-8")
    )
    assert data["success_count"] == 1
    assert data["failed_count"] == 1
    assert data["exit_code"] == 2  # runtime error (FileNotFoundError, not ConfigError)
    task_by_name = {t["name"]: t for t in data["tasks"]}
    assert task_by_name["t_ok"]["status"] == "success"
    assert task_by_name["t_bad_file"]["status"] == "failed"
    assert "error" in task_by_name["t_bad_file"]
