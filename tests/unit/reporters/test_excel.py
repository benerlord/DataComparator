import pandas as pd
from openpyxl import load_workbook
from datacompare.reporters.excel import ExcelReporter
from datacompare.engine.result import CompareResult


def _sample():
    return CompareResult(
        task_name="Sales", left_name="l", right_name="r",
        left_total=10, right_total=10,
        matched_rows=8, identical_rows=7, diff_rows=1,
        left_only=1, right_only=1,
        diff_details=pd.DataFrame([{"order_id": "A1", "field": "amount",
                                    "left_value": "1", "right_value": "2",
                                    "diff_type": "value_mismatch"}]),
        left_only_rows=pd.DataFrame([{"order_id": "X1", "amount": "1"}]),
        right_only_rows=pd.DataFrame([{"order_id": "Y1", "amount": "2"}]),
        engine_used="memory", duration_seconds=0.5, errors=[],
    )


def test_excel_writes_multi_sheet(tmp_path):
    p = ExcelReporter({"highlight_diff_cells": True}, tmp_path).render(_sample())
    assert p.exists()
    wb = load_workbook(p)
    assert set(wb.sheetnames) == {"摘要", "字段差异", "左侧独有", "右侧独有"}


def test_excel_summary_contains_metrics(tmp_path):
    p = ExcelReporter({"highlight_diff_cells": False}, tmp_path).render(_sample())
    wb = load_workbook(p)
    ws = wb["摘要"]
    values = [str(row[0].value) for row in ws.iter_rows()]
    assert any("Sales" in v for v in values)
