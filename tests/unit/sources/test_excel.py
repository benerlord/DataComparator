import pytest
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from datacompare.sources.excel import ExcelSource
from datacompare.config.models import ExcelSourceConfig, SheetSelector
from datacompare.config.errors import ConfigError

FIXTURES = Path(__file__).parent.parent.parent / "fixtures" / "excel"


@pytest.fixture(scope="module", autouse=True)
def _make_fixtures():
    FIXTURES.mkdir(parents=True, exist_ok=True)
    # simple.xlsx: single sheet, header row 1
    wb = Workbook()
    ws = wb.active
    ws.append(["order_id", "amount", "region"])
    ws.append(["A001", "100.50", "North"])
    ws.append(["A002", "200.00", "South"])
    wb.save(FIXTURES / "simple.xlsx")
    # multi_sheet.xlsx: two sheets with same header
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "North"
    ws1.append(["order_id", "amount"])
    ws1.append(["A001", "100"])
    ws2 = wb.create_sheet("South")
    ws2.append(["order_id", "amount"])
    ws2.append(["B001", "200"])
    ws2.append(["B002", "300"])
    wb.save(FIXTURES / "multi_sheet.xlsx")
    # header_row2.xlsx: title in row 1, headers in row 2
    wb = Workbook()
    ws = wb.active
    ws.append(["Report", None])
    ws.append(["order_id", "amount"])
    ws.append(["X1", "50"])
    wb.save(FIXTURES / "header_row2.xlsx")
    yield


def test_columns_from_first_sheet():
    cfg = ExcelSourceConfig(path=str(FIXTURES / "simple.xlsx"))
    src = ExcelSource(cfg)
    assert src.columns() == ["order_id", "amount", "region"]
    src.close()


def test_estimated_rows():
    cfg = ExcelSourceConfig(path=str(FIXTURES / "simple.xlsx"))
    src = ExcelSource(cfg)
    assert src.estimated_rows() == 2
    src.close()


def test_read_returns_strings():
    cfg = ExcelSourceConfig(path=str(FIXTURES / "simple.xlsx"))
    src = ExcelSource(cfg)
    chunks = list(src.read())
    assert len(chunks) == 1
    df = chunks[0]
    assert df.iloc[0]["order_id"] == "A001"
    assert df.iloc[0]["amount"] == "100.50"
    assert all(df.dtypes == "object")
    src.close()


def test_multi_sheet_by_name_concat():
    cfg = ExcelSourceConfig(
        path=str(FIXTURES / "multi_sheet.xlsx"),
        sheets=[SheetSelector(name="North"), SheetSelector(name="South")],
    )
    src = ExcelSource(cfg)
    df = pd.concat(src.read())
    assert len(df) == 3
    assert "__sheet__" in df.columns
    assert set(df["__sheet__"].unique()) == {"North", "South"}
    src.close()


def test_header_row_configurable():
    cfg = ExcelSourceConfig(path=str(FIXTURES / "header_row2.xlsx"), header_row=2)
    src = ExcelSource(cfg)
    assert src.columns() == ["order_id", "amount"]
    df = pd.concat(src.read())
    assert df.iloc[0]["order_id"] == "X1"
    src.close()


def test_multi_sheet_header_mismatch_raises(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "A"
    ws1.append(["id", "amount"])
    ws1.append(["1", "100"])
    ws2 = wb.create_sheet("B")
    ws2.append(["id", "value"])   # mismatched header
    ws2.append(["2", "200"])
    p = tmp_path / "mismatch.xlsx"
    wb.save(p)
    cfg = ExcelSourceConfig(
        path=str(p),
        sheets=[SheetSelector(name="A"), SheetSelector(name="B")],
    )
    src = ExcelSource(cfg)
    with pytest.raises(Exception, match="header"):
        src.columns()
    src.close()


@pytest.fixture
def dated_sheets_xlsx(tmp_path):
    """3-sheet Excel with date-suffixed names for name_regex tests."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "物理主机_2026_07"
    ws1.append(["id", "name"])
    ws1.append(["p1", "host-1"])
    ws2 = wb.create_sheet("云主机_2026_07")
    ws2.append(["id", "name"])
    ws2.append(["v1", "vm-1"])
    ws3 = wb.create_sheet("存储_2026_07")
    ws3.append(["id", "name"])
    ws3.append(["s1", "disk-1"])
    p = tmp_path / "dated.xlsx"
    wb.save(p)
    return p


def test_excel_source_name_regex_unique_match(dated_sheets_xlsx):
    """正则唯一命中 → 返回该 sheet 的数据。"""
    cfg = ExcelSourceConfig(
        path=str(dated_sheets_xlsx),
        sheets=[SheetSelector(name_regex=r"^物理主机_\d{4}_\d{2}$")],
    )
    src = ExcelSource(cfg)
    df = pd.concat(src.read())
    assert len(df) == 1
    assert df.iloc[0]["id"] == "p1"
    assert set(df["__sheet__"].unique()) == {"物理主机_2026_07"}
    src.close()


def test_excel_source_name_regex_zero_match_raises(dated_sheets_xlsx):
    """0 命中 → ConfigError，suggestion 列出可用 sheet。"""
    cfg = ExcelSourceConfig(
        path=str(dated_sheets_xlsx),
        sheets=[SheetSelector(name_regex=r"^数据库_\d{4}_\d{2}$")],
    )
    src = ExcelSource(cfg)
    with pytest.raises(ConfigError) as excinfo:
        src.columns()
    msg = str(excinfo.value)
    assert "matched no sheets" in msg
    assert "物理主机_2026_07" in msg
    src.close()


def test_excel_source_name_regex_multi_match_raises(dated_sheets_xlsx):
    """≥2 命中 → ConfigError，message 列出所有命中项。"""
    cfg = ExcelSourceConfig(
        path=str(dated_sheets_xlsx),
        sheets=[SheetSelector(name_regex=r".+_2026_07")],   # 3 sheets 都命中
    )
    src = ExcelSource(cfg)
    with pytest.raises(ConfigError) as excinfo:
        src.columns()
    msg = str(excinfo.value)
    assert "matched 3 sheets" in msg
    assert "物理主机_2026_07" in msg
    assert "云主机_2026_07" in msg
    assert "存储_2026_07" in msg
    src.close()


def test_excel_source_name_regex_case_insensitive_via_flag(tmp_path):
    """(?i) inline flag 让匹配大小写不敏感。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "PHYSICAL_HOST"
    ws.append(["id", "name"])
    ws.append(["1", "a"])
    p = tmp_path / "upper.xlsx"
    wb.save(p)

    cfg = ExcelSourceConfig(
        path=str(p),
        sheets=[SheetSelector(name_regex="(?i)^physical_host$")],
    )
    src = ExcelSource(cfg)
    df = pd.concat(src.read())
    assert set(df["__sheet__"].unique()) == {"PHYSICAL_HOST"}
    src.close()


def test_excel_source_mixed_selectors(dated_sheets_xlsx):
    """一个 sheets 列表里 name + name_regex + index 三种混用。"""
    cfg = ExcelSourceConfig(
        path=str(dated_sheets_xlsx),
        sheets=[
            SheetSelector(name="云主机_2026_07"),
            SheetSelector(name_regex=r"^物理主机_\d{4}_\d{2}$"),
            SheetSelector(index=2),   # 存储_2026_07
        ],
    )
    src = ExcelSource(cfg)
    df = pd.concat(src.read())
    assert len(df) == 3
    assert set(df["__sheet__"].unique()) == {
        "云主机_2026_07", "物理主机_2026_07", "存储_2026_07",
    }
    src.close()
