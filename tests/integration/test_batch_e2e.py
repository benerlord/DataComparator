"""End-to-end batch tests: heterogeneous right sides, mixed pass/fail, cross-sheet."""
import json
from pathlib import Path
import httpx
import respx
import yaml
from openpyxl import Workbook
from typer.testing import CliRunner
from datacompare.cli import app


runner = CliRunner()


def _make_xlsx(path: Path, sheets: dict[str, list[list]]):
    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "_placeholder"
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    if "_placeholder" in wb.sheetnames:
        del wb["_placeholder"]
    wb.save(path)


@respx.mock
def test_batch_scenario_g_heterogeneous_success(tmp_path):
    """Scenario G: 3 sub-tasks — Excel→Excel(sheet), Excel→Excel(other file), Excel→API."""
    _make_xlsx(tmp_path / "manage.xlsx", {
        "PHYSICAL": [["id", "name"], ["p1", "host-1"], ["p2", "host-2"]],
        "VM": [["id", "name"], ["v1", "vm-1"], ["v2", "vm-2"]],
        "API_DATA": [["id", "value"], ["a1", "10"], ["a2", "20"]],
    })
    _make_xlsx(tmp_path / "snapshot.xlsx", {
        "PHYSICAL": [["id", "name"], ["p1", "host-1"], ["p2", "host-2"]],
    })
    _make_xlsx(tmp_path / "vm_ref.xlsx", {
        "VM": [["id", "name"], ["v1", "vm-1"], ["v2", "vm-2"]],
    })
    respx.get("http://api.test/v1/data", params={"page": "1", "size": "100"}).mock(
        return_value=httpx.Response(200, json={"data": {"list": [
            {"id": "a1", "value": "10"}, {"id": "a2", "value": "20"},
        ]}})
    )
    respx.get("http://api.test/v1/data", params={"page": "2", "size": "100"}).mock(
        return_value=httpx.Response(200, json={"data": {"list": []}})
    )
    conns = tmp_path / "conns.yaml"
    conns.write_text("""
api_svc:
  type: api
  base_url: http://api.test
""", encoding="utf-8")
    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: hetero
sources:
  left: {{type: excel, path: {tmp_path}/manage.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: cross_sheet
    sources:
      left: {{sheets: [{{name: PHYSICAL}}]}}
      right: {{type: excel, path: {tmp_path}/snapshot.xlsx, sheets: [{{name: PHYSICAL}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: name, right: name}}]}}
  - name: vs_another_excel
    sources:
      left: {{sheets: [{{name: VM}}]}}
      right: {{type: excel, path: {tmp_path}/vm_ref.xlsx, sheets: [{{name: VM}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: name, right: name}}]}}
  - name: vs_api
    sources:
      left: {{sheets: [{{name: API_DATA}}]}}
      right:
        type: api
        connection: api_svc
        url: /v1/data
        pagination: {{type: page, page_param: page, size_param: size, size: 100}}
        data_path: $.data.list[*]
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: value, right: value}}]}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(conns)])
    assert result.exit_code == 0, result.output
    for sub in ["cross_sheet", "vs_another_excel", "vs_api"]:
        assert (tmp_path / "reports" / sub / "report.json").exists()
    batch_log = tmp_path / "reports" / "batch.log"
    entries = [json.loads(l) for l in batch_log.read_text(encoding="utf-8").splitlines() if l.strip()]
    task_ends = [e for e in entries if e["event"] == "task_end"]
    assert len(task_ends) == 3
    assert all(e["status"] == "success" for e in task_ends)


@respx.mock
def test_batch_scenario_h_mixed_failure_continue(tmp_path):
    """Scenario H: 3 sub-tasks — 1 success + 1 missing file + 1 API 500."""
    _make_xlsx(tmp_path / "manage.xlsx", {
        "OK_SHEET": [["id"], ["x1"]],
        "MISSING_RIGHT": [["id"], ["y1"]],
        "API_500": [["id"], ["z1"]],
    })
    _make_xlsx(tmp_path / "ok_right.xlsx", {
        "OK_SHEET": [["id"], ["x1"]],
    })
    respx.get("http://api.test/broken").mock(return_value=httpx.Response(500))
    conns = tmp_path / "conns.yaml"
    conns.write_text("api_svc: {type: api, base_url: http://api.test}\n", encoding="utf-8")
    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: mixed
on_error: continue
sources:
  left: {{type: excel, path: {tmp_path}/manage.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: ok_task
    sources:
      left: {{sheets: [{{name: OK_SHEET}}]}}
      right: {{type: excel, path: {tmp_path}/ok_right.xlsx, sheets: [{{name: OK_SHEET}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
  - name: file_missing
    sources:
      left: {{sheets: [{{name: MISSING_RIGHT}}]}}
      right: {{type: excel, path: {tmp_path}/DOES_NOT_EXIST.xlsx}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
  - name: api_broken
    sources:
      left: {{sheets: [{{name: API_500}}]}}
      right:
        type: api
        connection: api_svc
        url: /broken
        pagination: {{type: page, page_param: page, size_param: size, size: 100}}
        data_path: $.data
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(conns)])
    assert result.exit_code == 2
    assert (tmp_path / "reports" / "ok_task" / "report.json").exists()
    batch_log = tmp_path / "reports" / "batch.log"
    entries = [json.loads(l) for l in batch_log.read_text(encoding="utf-8").splitlines() if l.strip()]
    task_ends = {e["task_name"]: e for e in entries if e["event"] == "task_end"}
    assert task_ends["ok_task"]["status"] == "success"
    assert task_ends["file_missing"]["status"] == "failed"
    assert task_ends["api_broken"]["status"] == "failed"


def test_batch_scenario_i_same_excel_cross_sheet(tmp_path):
    """Scenario I: sub-task compares two sheets from the same Excel file."""
    _make_xlsx(tmp_path / "same.xlsx", {
        "LEFT_SHEET": [["id", "v"], ["a", "1"], ["b", "2"]],
        "RIGHT_SHEET": [["id", "v"], ["a", "1"], ["b", "2"]],
    })
    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: same_file_cross_sheet
sources:
  left: {{type: excel, path: {tmp_path}/same.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: sheet_a_vs_sheet_b
    sources:
      left: {{sheets: [{{name: LEFT_SHEET}}]}}
      right: {{type: excel, path: {tmp_path}/same.xlsx, sheets: [{{name: RIGHT_SHEET}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: v, right: v}}]}}
""", encoding="utf-8")
    result = runner.invoke(app, ["run", str(task), "--connections", str(tmp_path / "none.yaml")])
    assert result.exit_code == 0, result.output
    report = json.loads((tmp_path / "reports" / "sheet_a_vs_sheet_b" / "report.json").read_text(encoding="utf-8"))
    assert report["summary"]["matched"] == 2


def test_batch_scenario_j_left_literal_asserts_right_column_value(tmp_path):
    """Scenario J: sub-task uses `left_literal` to assert that a right-side
    column always equals a fixed value for every matched row.

    Left Excel has only `id` column. Right Excel has `id` and `zone` columns
    where zone varies per row. The compare field `{left_literal: 'Azone',
    right: 'zone'}` should produce diffs for exactly the rows where right's
    `zone != 'Azone'`.
    """
    _make_xlsx(tmp_path / "left.xlsx", {
        "IDS": [["id"], ["r1"], ["r2"], ["r3"]],
    })
    _make_xlsx(tmp_path / "right.xlsx", {
        "ZONES": [
            ["id", "zone"],
            ["r1", "Azone"],   # matches literal
            ["r2", "Bzone"],   # diff
            ["r3", "Azone"],   # matches
        ],
    })
    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: literal_assertion
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: assert_zone_is_Azone
    sources:
      left: {{sheets: [{{name: IDS}}]}}
      right: {{type: excel, path: {tmp_path}/right.xlsx, sheets: [{{name: ZONES}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare:
      fields:
        - {{left_literal: "Azone", right: zone}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(tmp_path / "none.yaml")])
    assert result.exit_code == 0, result.output
    report = json.loads(
        (tmp_path / "reports" / "assert_zone_is_Azone" / "report.json").read_text(encoding="utf-8")
    )
    assert report["summary"]["matched"] == 3
    assert report["summary"]["identical"] == 2
    assert report["summary"]["diff"] == 1


def test_batch_scenario_k_key_alias_and_field_regex(tmp_path):
    """Scenario K: batch sub-task uses key alias + field regex to compare
    a compound right-side column against split left-side columns.

    Left: {id, name}. Right: {name} = "prefix@@id" pattern.
    Join on right's regex-extracted ID via alias=join_id (avoids name collision).
    Compare left.name against right.name regex-extracted prefix."""
    _make_xlsx(tmp_path / "left.xlsx", {
        "USERS": [["id", "name"], ["1", "Alice"], ["2", "Bob"], ["3", "Carol"]],
    })
    _make_xlsx(tmp_path / "right.xlsx", {
        "COMPOUND": [["name"], ["Alice@@1"], ["Diff@@2"], ["Carol@@3"]],
    })
    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: alias_and_field_regex_batch
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: users_vs_compound
    sources:
      left: {{sheets: [{{name: USERS}}]}}
      right: {{type: excel, path: {tmp_path}/right.xlsx, sheets: [{{name: COMPOUND}}]}}
    match:
      keys:
        - {{left: id, right: name, right_regex: '.*@@(.*)', alias: join_id}}
    compare:
      fields:
        - {{left: name, right: name, right_regex: '(.*)@@.*'}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(tmp_path / "none.yaml")])
    assert result.exit_code == 0, result.output
    report = json.loads(
        (tmp_path / "reports" / "users_vs_compound" / "report.json").read_text(encoding="utf-8")
    )
    assert report["summary"]["matched"] == 3
    assert report["summary"]["identical"] == 2
    assert report["summary"]["diff"] == 1


def test_batch_scenario_l_summary_report_with_failure(tmp_path):
    """Scenario L: batch with 1 success + 1 failed (bad file) + 1 skipped
    (fail_fast) produces batch_summary.{json,html} with all statuses reflected.
    """
    _make_xlsx(tmp_path / "left.xlsx", {
        "GOOD": [["id"], ["1"], ["2"]],
        "OTHER": [["id"], ["3"]],
    })
    _make_xlsx(tmp_path / "right.xlsx", {
        "GOOD": [["id"], ["1"], ["2"]],
    })
    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: scenario_l
on_error: fail_fast
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: ok_task
    sources:
      left: {{sheets: [{{name: GOOD}}]}}
      right: {{type: excel, path: {tmp_path}/right.xlsx, sheets: [{{name: GOOD}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
  - name: bad_file_task
    sources:
      left: {{sheets: [{{name: OTHER}}]}}
      right: {{type: excel, path: {tmp_path}/does_not_exist.xlsx, sheets: [{{name: X}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
  - name: skipped_after_failure
    sources:
      left: {{sheets: [{{name: GOOD}}]}}
      right: {{type: excel, path: {tmp_path}/right.xlsx, sheets: [{{name: GOOD}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: []}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(tmp_path / "none.yaml")])
    assert result.exit_code == 2

    summary_json = tmp_path / "reports" / "batch_summary.json"
    assert summary_json.exists()
    data = json.loads(summary_json.read_text(encoding="utf-8"))
    assert data["batch_name"] == "scenario_l"
    assert data["task_count"] == 3
    assert data["success_count"] == 1
    assert data["failed_count"] == 1
    assert data["skipped_count"] == 1
    assert data["exit_code"] == 2
    by_name = {t["name"]: t for t in data["tasks"]}
    assert by_name["ok_task"]["status"] == "success"
    assert "stats" in by_name["ok_task"]
    assert by_name["bad_file_task"]["status"] == "failed"
    assert "error" in by_name["bad_file_task"]
    assert by_name["skipped_after_failure"]["status"] == "skipped"

    summary_html = tmp_path / "reports" / "batch_summary.html"
    assert summary_html.exists()
    html_text = summary_html.read_text(encoding="utf-8")
    for name in ("ok_task", "bad_file_task", "skipped_after_failure"):
        assert name in html_text
    assert "✓" in html_text
    assert "✗" in html_text
    assert 'href="ok_task/report.html"' in html_text


def test_batch_scenario_m_field_missing_soft_fail(tmp_path):
    """v0.8 Scenario M: 3-sub-task batch
      - task1: 正常成功
      - task2: 单侧 field 缺列 → 从 v0.7 的 failed 变成 success + field_missing 汇总
      - task3: key 缺列 → 仍 failed（key 硬失败路径不变）
    """
    _make_xlsx(tmp_path / "left.xlsx", {
        "T1": [["id", "name"], ["1", "a"], ["2", "b"]],
        # T2: 左侧只有 vmemory（打字错误 vmemorys 在左侧不存在）
        "T2": [["id", "vmemory"], ["1", "16"], ["2", "32"]],
        # T3: 左侧缺 id 列，触发 key 硬失败
        "T3": [["name_only"], ["x"]],
    })
    _make_xlsx(tmp_path / "right.xlsx", {
        "T1": [["id", "name"], ["1", "a"], ["2", "b"]],
        # T2: 右侧有 vmemory 和 vmemorys
        "T2": [["id", "vmemory", "vmemorys"], ["1", "16", "16GB"], ["2", "32", "32GB"]],
        "T3": [["id", "name_only"], ["1", "x"]],
    })

    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: scenario_m
on_error: continue
sources:
  left: {{type: excel, path: {tmp_path}/left.xlsx}}
  right: {{type: excel, path: {tmp_path}/right.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [html, json]
tasks:
  - name: task1_ok
    sources:
      left: {{sheets: [{{name: T1}}]}}
      right: {{sheets: [{{name: T1}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: name, right: name}}]}}
  - name: task2_field_missing
    sources:
      left: {{sheets: [{{name: T2}}]}}
      right: {{sheets: [{{name: T2}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare:
      fields:
        - {{left: vmemory, right: vmemory}}
        - {{left: vmemorys, right: vmemorys}}
  - name: task3_key_missing
    sources:
      left: {{sheets: [{{name: T3}}]}}
      right: {{sheets: [{{name: T3}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: name_only, right: name_only}}]}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(tmp_path / "none.yaml")])
    # task3 fails with ConfigError → exit code 1
    assert result.exit_code == 1, f"stdout={result.output}"

    summary_json = tmp_path / "reports" / "batch_summary.json"
    assert summary_json.exists()
    data = json.loads(summary_json.read_text(encoding="utf-8"))
    assert data["batch_name"] == "scenario_m"
    assert data["task_count"] == 3
    assert data["success_count"] == 2   # task1 + task2
    assert data["failed_count"] == 1    # task3
    assert data["skipped_count"] == 0
    by_name = {t["name"]: t for t in data["tasks"]}
    assert by_name["task1_ok"]["status"] == "success"
    assert by_name["task2_field_missing"]["status"] == "success"
    # task2 至少产生一条 field_missing 汇总 diff（vmemorys 左侧缺）
    assert by_name["task2_field_missing"]["stats"]["diff"] >= 1
    assert by_name["task3_key_missing"]["status"] == "failed"

    # task2 生成完整报告（v0.7 之前空目录）
    task2_dir = tmp_path / "reports" / "task2_field_missing"
    assert (task2_dir / "report.html").exists()
    assert (task2_dir / "report.json").exists()

    # task2 report.json 里应能看到 field_missing 类型的 diff 记录
    task2_report = json.loads((task2_dir / "report.json").read_text(encoding="utf-8"))
    diff_types = {d.get("diff_type") for d in task2_report.get("diff_details", [])}
    assert "field_missing" in diff_types

    # batch_summary.html 里 task2 显示 ✓ 而不是 ✗
    summary_html = tmp_path / "reports" / "batch_summary.html"
    html_text = summary_html.read_text(encoding="utf-8")
    assert "task2_field_missing" in html_text
    # 找到 task2 附近应有成功标记
    idx = html_text.find("task2_field_missing")
    surrounding = html_text[max(0, idx - 400):idx + 400]
    assert "✓" in surrounding


def test_batch_scenario_n_sheet_name_regex(tmp_path):
    """v0.9 Scenario N: 批次模式下用 name_regex 定位一张日期戳变名的 sheet。
    - Excel 有 3 张 sheet：物理主机_2026_07 / 云主机_2026_07 / 存储_2026_07
    - batch.yaml 用 name_regex "^物理主机_\\d{4}_\\d{2}$" 定位第一张
    - 断言 sub-task 成功、__sheet__ 列值正确
    """
    _make_xlsx(tmp_path / "manage.xlsx", {
        "物理主机_2026_07": [["id", "name"], ["p1", "host-1"], ["p2", "host-2"]],
        "云主机_2026_07": [["id", "name"], ["v1", "vm-1"]],
        "存储_2026_07": [["id", "name"], ["s1", "disk-1"]],
    })
    _make_xlsx(tmp_path / "snapshot.xlsx", {
        "PHYSICAL": [["id", "name"], ["p1", "host-1"], ["p2", "host-2"]],
    })

    task = tmp_path / "batch.yaml"
    task.write_text(f"""
name: scenario_n
sources:
  left: {{type: excel, path: {tmp_path}/manage.xlsx}}
output:
  dir: {tmp_path}/reports
  formats: [json]
tasks:
  - name: physical_via_regex
    sources:
      left: {{sheets: [{{name_regex: "^物理主机_\\\\d{{4}}_\\\\d{{2}}$"}}]}}
      right: {{type: excel, path: {tmp_path}/snapshot.xlsx, sheets: [{{name: PHYSICAL}}]}}
    match: {{keys: [{{left: id, right: id}}]}}
    compare: {{fields: [{{left: name, right: name}}]}}
""", encoding="utf-8")

    result = runner.invoke(app, ["run", str(task), "--connections", str(tmp_path / "none.yaml")])
    assert result.exit_code == 0, f"stdout={result.output}"

    summary_json = tmp_path / "reports" / "batch_summary.json"
    assert summary_json.exists()
    data = json.loads(summary_json.read_text(encoding="utf-8"))
    assert data["success_count"] == 1
    task_entry = data["tasks"][0]
    assert task_entry["name"] == "physical_via_regex"
    assert task_entry["status"] == "success"
    assert task_entry["stats"]["matched"] == 2
    assert task_entry["stats"]["diff"] == 0

    report_json = tmp_path / "reports" / "physical_via_regex" / "report.json"
    assert report_json.exists()
