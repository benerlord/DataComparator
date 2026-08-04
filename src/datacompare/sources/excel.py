"""Excel source using openpyxl in read-only mode."""
from __future__ import annotations
from typing import Iterator
import re
import structlog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from .base import DataSource
from .registry import register_source
from datacompare.config.models import ExcelSourceConfig, SheetSelector
from datacompare.config.errors import ConfigError

_log = structlog.get_logger(__name__)


@register_source("excel")
class ExcelSource(DataSource):
    def __init__(self, config: ExcelSourceConfig, name: str = ""):
        self.config = config
        self.name = name or f"excel:{config.path}"
        self._wb: Workbook | None = None

    def _open(self) -> Workbook:
        if self._wb is None:
            self._wb = load_workbook(self.config.path, read_only=True, data_only=True)
        return self._wb

    def _selected_sheet_names(self) -> list[str]:
        wb = self._open()
        result: list[str] = []
        for sel in self.config.sheets:
            if sel.name is not None:
                if sel.name not in wb.sheetnames:
                    raise ConfigError(
                        f"sheet '{sel.name}' not found",
                        suggestion=f"available: {wb.sheetnames}",
                    )
                result.append(sel.name)
            elif sel.index is not None:
                if sel.index >= len(wb.sheetnames):
                    raise ConfigError(f"sheet index {sel.index} out of range")
                result.append(wb.sheetnames[sel.index])
            elif sel.name_regex is not None:
                pattern = re.compile(sel.name_regex)
                matches = [n for n in wb.sheetnames if pattern.fullmatch(n)]
                if len(matches) == 0:
                    raise ConfigError(
                        f"name_regex '{sel.name_regex}' matched no sheets",
                        path="sources.sheets",
                        suggestion=f"available: {wb.sheetnames}",
                    )
                if len(matches) > 1:
                    raise ConfigError(
                        f"name_regex '{sel.name_regex}' matched {len(matches)} sheets: {matches}",
                        path="sources.sheets",
                        suggestion="tighten the pattern to match exactly one sheet",
                    )
                _log.info(
                    "sheet_regex_resolved",
                    regex=sel.name_regex,
                    resolved_to=matches[0],
                )
                result.append(matches[0])
            else:
                # 不可达：validator 已强制三选一
                raise ConfigError("SheetSelector must have name, index, or name_regex")
        return result

    def _sheet_header(self, sheet_name: str) -> list[str]:
        wb = self._open()
        ws = wb[sheet_name]
        header_row_idx = self.config.header_row
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == header_row_idx:
                return [str(c) if c is not None else "" for c in row if c is not None]
        return []

    def columns(self) -> list[str]:
        sheets = self._selected_sheet_names()
        first_header = self._sheet_header(sheets[0])
        for name in sheets[1:]:
            other = self._sheet_header(name)
            if other != first_header:
                raise ConfigError(
                    f"sheet header mismatch: '{sheets[0]}' vs '{name}'",
                    suggestion=f"headers: {first_header} vs {other}",
                )
        return first_header

    def estimated_rows(self) -> int | None:
        wb = self._open()
        total = 0
        for name in self._selected_sheet_names():
            ws = wb[name]
            # openpyxl read_only max_row is exact
            total += max(0, (ws.max_row or 0) - self.config.header_row)
        return total

    def read(self, chunk_size: int = 100_000) -> Iterator[pd.DataFrame]:
        header = self.columns()
        wb = self._open()
        buffer: list[dict[str, str | None]] = []
        for sheet_name in self._selected_sheet_names():
            ws = wb[sheet_name]
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i <= self.config.header_row:
                    continue
                if all(v is None for v in row):
                    continue
                record: dict[str, str | None] = {"__sheet__": sheet_name}
                for col_name, cell in zip(header, row):
                    if cell is None:
                        record[col_name] = None
                    elif self.config.force_string:
                        record[col_name] = str(cell)
                    else:
                        record[col_name] = cell
                buffer.append(record)
                if len(buffer) >= chunk_size:
                    yield pd.DataFrame(buffer).astype(object)
                    buffer = []
        if buffer:
            yield pd.DataFrame(buffer).astype(object)

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None
